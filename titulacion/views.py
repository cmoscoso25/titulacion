import json

from django.contrib import messages
from django.db import IntegrityError, transaction
from django.db.models import Count, Max, Min, Q
from django.db.models.functions import TruncHour
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from .generador_qr import generar_qr_estudiante, generar_qr_invitacion
from .models import (
    AreaAcademica,
    BloqueCeremonia,
    CambioCeremonia,
    Ceremonia,
    EstudianteTitulado,
    Invitacion,
    PlanEstudio,
    RegistroIngreso,
    generar_codigo_qr_estudiante,
    generar_codigo_qr_invitacion,
)
from .permisos import es_admin_titulacion, es_admision, es_curricular, es_entrega
from .procesador_excel import crear_invitaciones_para_estudiante, procesar_excel_titulados


DIRECCION_CEREMONIA = "Avenida Chile 1108"


DIAS_SEMANA = {
    0: "LUNES",
    1: "MARTES",
    2: "MIÉRCOLES",
    3: "JUEVES",
    4: "VIERNES",
    5: "SÁBADO",
    6: "DOMINGO",
}


MESES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


def formatear_hora(hora):
    if not hora:
        return ""
    return hora.strftime("%H:%M")


def datos_fecha_bloque(bloque):
    if not bloque or not bloque.fecha:
        return {
            "fecha_es": "",
            "hora_es": "",
            "dia_semana": "",
            "numero_dia": "",
            "mes_es": "",
        }

    fecha = bloque.fecha

    return {
        "fecha_es": f"{fecha.day:02d} de {MESES[fecha.month].lower()} de {fecha.year}",
        "hora_es": formatear_hora(bloque.hora_inicio),
        "dia_semana": DIAS_SEMANA[fecha.weekday()],
        "numero_dia": f"{fecha.day:02d}",
        "mes_es": MESES[fecha.month],
    }


def obtener_area_estudiante(estudiante):
    if estudiante and estudiante.plan_estudio and estudiante.plan_estudio.area:
        return estudiante.plan_estudio.area.nombre
    return ""


def obtener_plan_estudiante(estudiante):
    if estudiante and estudiante.plan_estudio:
        return estudiante.plan_estudio.nombre
    return ""


def obtener_bloque_estudiante(estudiante):
    if estudiante and estudiante.bloque_ceremonia:
        return estudiante.bloque_ceremonia.nombre
    return ""


def obtener_ceremonia_activa():
    return BloqueCeremonia.objects.filter(
        estado_registro="ABIERTA"
    ).select_related(
        "ceremonia"
    ).first()


def bloque_esta_cerrado(bloque):
    if not bloque:
        return False
    return bloque.estado_registro == "CERRADA"


def asegurar_qr_estudiante(estudiante):
    if not estudiante.codigo_qr_estudiante or not estudiante.imagen_qr_estudiante:
        generar_qr_estudiante(estudiante)

    for invitacion in estudiante.invitaciones.all():
        if not invitacion.codigo_qr or not invitacion.imagen_qr:
            generar_qr_invitacion(invitacion)


def regenerar_qr_estudiante_e_invitaciones(estudiante):
    if estudiante.imagen_qr_estudiante:
        estudiante.imagen_qr_estudiante.delete(save=False)

    estudiante.codigo_qr_estudiante = generar_codigo_qr_estudiante()
    estudiante.imagen_qr_estudiante = None
    estudiante.save()

    generar_qr_estudiante(estudiante)

    for invitacion in estudiante.invitaciones.all():
        if invitacion.imagen_qr:
            invitacion.imagen_qr.delete(save=False)

        invitacion.codigo_qr = generar_codigo_qr_invitacion()
        invitacion.imagen_qr = None
        invitacion.usada = False
        invitacion.fecha_uso = None
        invitacion.save()

        generar_qr_invitacion(invitacion)


def aplicar_filtros_estudiantes(queryset, request):
    busqueda = request.GET.get("q", "").strip()
    bloque_id = request.GET.get("bloque", "").strip()
    ceremonia_id = request.GET.get("ceremonia", "").strip()
    area = request.GET.get("area", "").strip()
    plan = request.GET.get("plan", "").strip()
    estado = request.GET.get("estado", "").strip()

    if busqueda:
        queryset = queryset.filter(
            Q(rut__icontains=busqueda) |
            Q(nombre_completo__icontains=busqueda) |
            Q(plan_estudio__nombre__icontains=busqueda) |
            Q(plan_estudio__area__nombre__icontains=busqueda) |
            Q(bloque_ceremonia__nombre__icontains=busqueda)
        )

    if bloque_id:
        queryset = queryset.filter(bloque_ceremonia_id=bloque_id)

    if ceremonia_id:
        queryset = queryset.filter(bloque_ceremonia__ceremonia_id=ceremonia_id)

    if area:
        queryset = queryset.filter(plan_estudio__area__nombre=area)

    if plan:
        queryset = queryset.filter(plan_estudio__nombre=plan)

    if estado == "presentes":
        queryset = queryset.filter(ingreso_confirmado=True)

    if estado == "pendientes":
        queryset = queryset.filter(ingreso_confirmado=False)

    if estado == "atrasados":
        queryset = queryset.filter(
            registros__resultado="ATRASADO"
        ).distinct()

    return queryset


def aplicar_filtros_registros(queryset, request):
    busqueda = request.GET.get("q", "").strip()
    bloque_id = request.GET.get("bloque", "").strip()
    ceremonia_id = request.GET.get("ceremonia", "").strip()
    area = request.GET.get("area", "").strip()
    plan = request.GET.get("plan", "").strip()
    tipo_acceso = request.GET.get("tipo_acceso", "").strip()
    resultado = request.GET.get("resultado", "").strip()

    if busqueda:
        queryset = queryset.filter(
            Q(estudiante__rut__icontains=busqueda) |
            Q(estudiante__nombre_completo__icontains=busqueda) |
            Q(estudiante__plan_estudio__nombre__icontains=busqueda) |
            Q(estudiante__plan_estudio__area__nombre__icontains=busqueda) |
            Q(estudiante__bloque_ceremonia__nombre__icontains=busqueda) |
            Q(observacion__icontains=busqueda)
        )

    if bloque_id:
        queryset = queryset.filter(estudiante__bloque_ceremonia_id=bloque_id)

    if ceremonia_id:
        queryset = queryset.filter(estudiante__bloque_ceremonia__ceremonia_id=ceremonia_id)

    if area:
        queryset = queryset.filter(estudiante__plan_estudio__area__nombre=area)

    if plan:
        queryset = queryset.filter(estudiante__plan_estudio__nombre=plan)

    if tipo_acceso:
        queryset = queryset.filter(tipo=tipo_acceso)

    if resultado:
        queryset = queryset.filter(resultado=resultado)

    return queryset


def inicio(request):
    total_estudiantes = EstudianteTitulado.objects.count()

    estudiantes_presentes = EstudianteTitulado.objects.filter(
        ingreso_confirmado=True
    ).count()

    invitados_presentes = Invitacion.objects.filter(
        usada=True
    ).count()

    total_atrasados = RegistroIngreso.objects.filter(
        resultado="ATRASADO"
    ).count()

    bloques = BloqueCeremonia.objects.select_related(
        "ceremonia"
    ).order_by(
        "fecha",
        "hora_inicio",
        "nombre"
    )

    u = request.user
    return render(
        request,
        "titulacion/inicio.html",
        {
            "total_estudiantes": total_estudiantes,
            "estudiantes_presentes": estudiantes_presentes,
            "invitados_presentes": invitados_presentes,
            "total_asistentes": estudiantes_presentes + invitados_presentes,
            "total_atrasados": total_atrasados,
            "bloques": bloques,
            "perm_admin":     es_admin_titulacion(u),
            "perm_admision":  es_admision(u),
            "perm_curricular": es_curricular(u),
            "perm_entrega":   es_entrega(u),
        }
    )


def cargar_excel(request):
    resumen = None

    if request.method == "POST":
        archivo_excel = request.FILES.get("archivo_excel")

        if not archivo_excel:
            messages.error(request, "Debes seleccionar un archivo Excel.")
            return redirect("titulacion:cargar_excel")

        try:
            resumen = procesar_excel_titulados(archivo_excel)

            messages.success(
                request,
                "Archivo Excel procesado correctamente."
            )

        except Exception as error:
            messages.error(
                request,
                f"Ocurrió un error al procesar el archivo: {str(error)}"
            )

    estudiantes_preview = EstudianteTitulado.objects.select_related(
        "bloque_ceremonia",
        "bloque_ceremonia__ceremonia",
        "plan_estudio",
        "plan_estudio__area",
    ).order_by("-id")[:30]

    return render(
        request,
        "titulacion/cargar_excel.html",
        {
            "resumen": resumen,
            "estudiantes_preview": estudiantes_preview,
            "estudiantes": estudiantes_preview,
            "total_estudiantes": EstudianteTitulado.objects.count(),
            "total_ceremonias": Ceremonia.objects.count(),
            "total_bloques": BloqueCeremonia.objects.count(),
            "total_areas": AreaAcademica.objects.count(),
            "total_planes": PlanEstudio.objects.count(),
            "total_invitaciones": Invitacion.objects.count(),
            "total_alertas": RegistroIngreso.objects.filter(
                resultado__in=["DENEGADO", "DUPLICADO", "OTRA_CEREMONIA"]
            ).count(),
        }
    )


def registro_ingreso(request):
    bloques = BloqueCeremonia.objects.select_related(
        "ceremonia"
    ).order_by(
        "fecha",
        "hora_inicio",
        "nombre"
    )

    ultimos_registros = RegistroIngreso.objects.select_related(
        "estudiante",
        "invitacion",
        "estudiante__bloque_ceremonia",
        "estudiante__plan_estudio",
        "estudiante__plan_estudio__area",
    ).order_by("-fecha_hora")[:20]

    return render(
        request,
        "titulacion/registro_ingreso.html",
        {
            "bloques": bloques,
            "ultimos_registros": ultimos_registros,
            "bloque_activo": obtener_ceremonia_activa(),
        }
    )


def ultimos_registros_ajax(request):
    registros = RegistroIngreso.objects.select_related(
        "estudiante",
    ).order_by("-fecha_hora")[:20]

    data = [
        {
            "fecha_hora": r.fecha_hora.strftime("%d/%m %H:%M"),
            "nombre": r.estudiante.nombre_completo if r.estudiante else "-",
            "tipo": r.tipo,
            "resultado": r.resultado,
        }
        for r in registros
    ]
    return JsonResponse({"registros": data})


@csrf_exempt
@require_POST
def validar_codigo_ingreso(request):

    codigo = request.POST.get("codigo", "").strip().upper()

    # =========================================================
    # NORMALIZACIÓN PARA LECTORES QR
    # Algunos lectores reemplazan "-" por "'" o caracteres raros
    # =========================================================
    codigo = (
        codigo
        .replace("'", "-")
        .replace("‘", "-")
        .replace("’", "-")
        .replace("`", "-")
        .replace("´", "-")
        .replace(" ", "")
        .replace("\n", "")
        .replace("\r", "")
    )

    if not codigo:
        return JsonResponse({
            "estado": "error",
            "ok": False,
            "titulo": "Código vacío",
            "mensaje": "No se recibió un código válido.",
        })

    bloque_activo = obtener_ceremonia_activa()

    if not bloque_activo:
        return JsonResponse({
            "estado": "rechazado",
            "ok": False,
            "tipo": "SIN_CEREMONIA_ACTIVA",
            "titulo": "Sin ceremonia activa",
            "mensaje": "Debe abrir una ceremonia antes de registrar ingresos.",
        })

    # =========================================================
    # VALIDAR ESTUDIANTE
    # =========================================================
    estudiante = EstudianteTitulado.objects.select_related(
        "bloque_ceremonia",
        "bloque_ceremonia__ceremonia",
        "plan_estudio",
        "plan_estudio__area",
    ).filter(
        codigo_qr_estudiante=codigo
    ).first()

    if estudiante:
        return registrar_ingreso_estudiante(
            estudiante=estudiante,
            bloque_activo=bloque_activo
        )

    # =========================================================
    # VALIDAR INVITACIÓN
    # =========================================================
    invitacion = Invitacion.objects.select_related(
        "estudiante",
        "estudiante__bloque_ceremonia",
        "estudiante__bloque_ceremonia__ceremonia",
        "estudiante__plan_estudio",
        "estudiante__plan_estudio__area",
    ).filter(
        codigo_qr=codigo
    ).first()

    if invitacion:
        return registrar_ingreso_invitado(
            invitacion=invitacion,
            bloque_activo=bloque_activo
        )

    # =========================================================
    # REGISTRO DE QR INVÁLIDO
    # =========================================================
    RegistroIngreso.objects.create(
        estudiante=None,
        invitacion=None,
        tipo="INVITADO",
        resultado="DENEGADO",
        observacion=f"Código inválido: {codigo}",
    )

    return JsonResponse({
        "estado": "error",
        "ok": False,
        "tipo": "INVALIDO",
        "titulo": "QR no encontrado",
        "mensaje": "El código escaneado no pertenece a esta ceremonia.",
    })


def registrar_ingreso_estudiante(estudiante, bloque_activo):
    bloque_estudiante = estudiante.bloque_ceremonia

    if bloque_estudiante.id != bloque_activo.id:
        RegistroIngreso.objects.create(
            estudiante=estudiante,
            invitacion=None,
            tipo="ESTUDIANTE",
            resultado="OTRA_CEREMONIA",
            observacion=(
                f"Intento de ingreso en {bloque_activo.nombre}. "
                f"El QR pertenece a {bloque_estudiante.nombre}."
            ),
        )

        return JsonResponse({
            "estado": "rechazado",
            "ok": False,
            "tipo": "OTRA_CEREMONIA",
            "titulo": "QR de otra ceremonia",
            "mensaje": f"Este QR pertenece a {bloque_estudiante.nombre}, no a {bloque_activo.nombre}.",
            "nombre": estudiante.nombre_completo,
            "rut": estudiante.rut,
            "area": obtener_area_estudiante(estudiante),
            "plan": obtener_plan_estudiante(estudiante),
            "bloque": obtener_bloque_estudiante(estudiante),
            "atrasado": False,
        })

    es_atrasado = bloque_esta_cerrado(bloque_estudiante)

    if estudiante.ingreso_confirmado:
        RegistroIngreso.objects.create(
            estudiante=estudiante,
            invitacion=None,
            tipo="ESTUDIANTE",
            resultado="DUPLICADO",
            observacion="Se intentó registrar nuevamente la tarjeta del estudiante.",
        )

        return JsonResponse({
            "estado": "rechazado",
            "ok": False,
            "tipo": "DUPLICADO",
            "titulo": "QR ya utilizado",
            "mensaje": "Este estudiante ya fue registrado anteriormente.",
            "nombre": estudiante.nombre_completo,
            "rut": estudiante.rut,
            "area": obtener_area_estudiante(estudiante),
            "plan": obtener_plan_estudiante(estudiante),
            "bloque": obtener_bloque_estudiante(estudiante),
            "atrasado": False,
        })

    estudiante.ingreso_confirmado = True
    estudiante.fecha_hora_ingreso = timezone.now()
    estudiante.save()

    resultado = "ATRASADO" if es_atrasado else "PERMITIDO"

    RegistroIngreso.objects.create(
        estudiante=estudiante,
        invitacion=None,
        tipo="ESTUDIANTE",
        resultado=resultado,
        observacion=(
            "Ingreso del estudiante registrado como atrasado."
            if es_atrasado
            else "Ingreso del estudiante registrado correctamente."
        ),
    )

    return JsonResponse({
        "estado": "permitido",
        "ok": True,
        "tipo": resultado,
        "titulo": "Ingreso atrasado registrado" if es_atrasado else "Ingreso autorizado",
        "mensaje": (
            "Estudiante registrado después del cierre de ingreso."
            if es_atrasado
            else "Estudiante registrado correctamente."
        ),
        "nombre": estudiante.nombre_completo,
        "rut": estudiante.rut,
        "area": obtener_area_estudiante(estudiante),
        "plan": obtener_plan_estudiante(estudiante),
        "bloque": obtener_bloque_estudiante(estudiante),
        "atrasado": es_atrasado,
    })


def registrar_ingreso_invitado(invitacion, bloque_activo):
    estudiante = invitacion.estudiante
    bloque_estudiante = estudiante.bloque_ceremonia

    if bloque_estudiante.id != bloque_activo.id:
        RegistroIngreso.objects.create(
            estudiante=estudiante,
            invitacion=invitacion,
            tipo="INVITADO",
            resultado="OTRA_CEREMONIA",
            observacion=(
                f"Intento de ingreso en {bloque_activo.nombre}. "
                f"La invitación pertenece a {bloque_estudiante.nombre}."
            ),
        )

        return JsonResponse({
            "estado": "rechazado",
            "ok": False,
            "tipo": "OTRA_CEREMONIA",
            "titulo": "Invitación de otra ceremonia",
            "mensaje": f"Esta invitación pertenece a {bloque_estudiante.nombre}, no a {bloque_activo.nombre}.",
            "nombre": estudiante.nombre_completo,
            "rut": estudiante.rut,
            "area": obtener_area_estudiante(estudiante),
            "plan": obtener_plan_estudiante(estudiante),
            "bloque": obtener_bloque_estudiante(estudiante),
            "invitacion": invitacion.numero_invitacion,
            "atrasado": False,
        })

    es_atrasado = bloque_esta_cerrado(bloque_estudiante)

    if invitacion.usada:
        RegistroIngreso.objects.create(
            estudiante=estudiante,
            invitacion=invitacion,
            tipo="INVITADO",
            resultado="DUPLICADO",
            observacion="Se intentó reutilizar una invitación.",
        )

        return JsonResponse({
            "estado": "rechazado",
            "ok": False,
            "tipo": "DUPLICADO",
            "titulo": "Invitación ya utilizada",
            "mensaje": "Esta invitación ya fue registrada anteriormente.",
            "nombre": estudiante.nombre_completo,
            "rut": estudiante.rut,
            "area": obtener_area_estudiante(estudiante),
            "plan": obtener_plan_estudiante(estudiante),
            "bloque": obtener_bloque_estudiante(estudiante),
            "invitacion": invitacion.numero_invitacion,
            "atrasado": False,
        })

    invitacion.usada = True
    invitacion.fecha_uso = timezone.now()
    invitacion.save()

    resultado = "ATRASADO" if es_atrasado else "PERMITIDO"

    RegistroIngreso.objects.create(
        estudiante=estudiante,
        invitacion=invitacion,
        tipo="INVITADO",
        resultado=resultado,
        observacion=(
            f"Ingreso atrasado registrado para invitación {invitacion.numero_invitacion}."
            if es_atrasado
            else f"Ingreso registrado para invitación {invitacion.numero_invitacion}."
        ),
    )

    return JsonResponse({
        "estado": "permitido",
        "ok": True,
        "tipo": resultado,
        "titulo": "Ingreso atrasado registrado" if es_atrasado else "Ingreso autorizado",
        "mensaje": (
            f"Invitación {invitacion.numero_invitacion} registrada después del cierre de ingreso."
            if es_atrasado
            else f"Invitación {invitacion.numero_invitacion} registrada correctamente."
        ),
        "nombre": estudiante.nombre_completo,
        "rut": estudiante.rut,
        "area": obtener_area_estudiante(estudiante),
        "plan": obtener_plan_estudiante(estudiante),
        "bloque": obtener_bloque_estudiante(estudiante),
        "invitacion": invitacion.numero_invitacion,
        "invitados_presentes": estudiante.invitaciones.filter(usada=True).count(),
        "atrasado": es_atrasado,
    })


def panel_control(request):
    bloques = BloqueCeremonia.objects.select_related(
        "ceremonia"
    ).order_by(
        "fecha",
        "hora_inicio",
        "nombre"
    )

    ceremonias = Ceremonia.objects.all().order_by("-anio", "nombre")
    ceremonia_principal = ceremonias.first()
    areas = AreaAcademica.objects.all().order_by("nombre")
    planes = PlanEstudio.objects.select_related("area").all().order_by("nombre")
    bloque_abierto = bloques.filter(estado_registro="ABIERTA").first()

    return render(
        request,
        "titulacion/panel_control.html",
        {
            "bloques": bloques,
            "ceremonias": ceremonias,
            "ceremonia_principal": ceremonia_principal,
            "areas": areas,
            "planes": planes,
            "bloque_abierto": bloque_abierto,
        }
    )


def datos_panel_control(request):
    estudiantes_base = EstudianteTitulado.objects.select_related(
        "bloque_ceremonia",
        "bloque_ceremonia__ceremonia",
        "plan_estudio",
        "plan_estudio__area",
    ).prefetch_related("invitaciones").all()

    estudiantes_filtrados = aplicar_filtros_estudiantes(estudiantes_base, request)

    registros_base = RegistroIngreso.objects.select_related(
        "estudiante",
        "invitacion",
        "estudiante__bloque_ceremonia",
        "estudiante__bloque_ceremonia__ceremonia",
        "estudiante__plan_estudio",
        "estudiante__plan_estudio__area",
    ).all()

    registros_filtrados = aplicar_filtros_registros(registros_base, request)

    estudiantes_totales = estudiantes_filtrados.count()
    estudiantes_presentes = estudiantes_filtrados.filter(ingreso_confirmado=True).count()
    estudiantes_pendientes = estudiantes_filtrados.filter(ingreso_confirmado=False).count()

    estudiantes_atrasados = RegistroIngreso.objects.filter(
        estudiante__in=estudiantes_filtrados,
        tipo="ESTUDIANTE",
        resultado="ATRASADO",
    ).values("estudiante_id").distinct().count()

    invitaciones_filtradas = Invitacion.objects.select_related(
        "estudiante",
        "estudiante__bloque_ceremonia",
        "estudiante__plan_estudio",
        "estudiante__plan_estudio__area",
    ).filter(estudiante__in=estudiantes_filtrados)

    invitados_presentes = invitaciones_filtradas.filter(usada=True).count()

    invitados_atrasados = RegistroIngreso.objects.filter(
        estudiante__in=estudiantes_filtrados,
        tipo="INVITADO",
        resultado="ATRASADO",
    ).count()

    invitaciones_totales = invitaciones_filtradas.count()
    invitaciones_pendientes = invitaciones_totales - invitados_presentes
    total_asistentes = estudiantes_presentes + invitados_presentes
    total_atrasados = estudiantes_atrasados + invitados_atrasados

    alertas_qr = registros_filtrados.filter(
        resultado__in=["DENEGADO", "DUPLICADO", "OTRA_CEREMONIA"]
    ).count()

    porcentaje_estudiantes = 0
    if estudiantes_totales > 0:
        porcentaje_estudiantes = round(
            (estudiantes_presentes / estudiantes_totales) * 100, 1
        )

    porcentaje_invitaciones = 0
    if invitaciones_totales > 0:
        porcentaje_invitaciones = round(
            (invitados_presentes / invitaciones_totales) * 100, 1
        )

    avance_por_plan = []

    planes = estudiantes_filtrados.values(
        "plan_estudio__nombre"
    ).annotate(
        total=Count("id"),
        presentes=Count("id", filter=Q(ingreso_confirmado=True)),
    ).order_by("plan_estudio__nombre")

    for item in planes:
        total = item["total"]
        presentes = item["presentes"]

        porcentaje = 0
        if total > 0:
            porcentaje = round((presentes / total) * 100, 1)

        atrasados_plan = RegistroIngreso.objects.filter(
            estudiante__in=estudiantes_filtrados.filter(
                plan_estudio__nombre=item["plan_estudio__nombre"]
            ),
            resultado="ATRASADO",
        ).count()

        avance_por_plan.append({
            "plan": item["plan_estudio__nombre"] or "Sin plan informado",
            "total": total,
            "presentes": presentes,
            "pendientes": total - presentes,
            "atrasados": atrasados_plan,
            "porcentaje": porcentaje,
        })

    seguimiento_estudiantes = []

    for estudiante in estudiantes_filtrados.order_by(
        "bloque_ceremonia__fecha",
        "bloque_ceremonia__hora_inicio",
        "nombre_completo"
    )[:300]:
        invitaciones = list(
            estudiante.invitaciones.all().order_by("numero_invitacion")
        )

        invitacion_1 = invitaciones[0] if len(invitaciones) >= 1 else None
        invitacion_2 = invitaciones[1] if len(invitaciones) >= 2 else None

        estudiante_atrasado = RegistroIngreso.objects.filter(
            estudiante=estudiante,
            tipo="ESTUDIANTE",
            resultado="ATRASADO",
        ).exists()

        invitacion_1_atrasada = False
        invitacion_2_atrasada = False

        if invitacion_1:
            invitacion_1_atrasada = RegistroIngreso.objects.filter(
                invitacion=invitacion_1,
                resultado="ATRASADO",
            ).exists()

        if invitacion_2:
            invitacion_2_atrasada = RegistroIngreso.objects.filter(
                invitacion=invitacion_2,
                resultado="ATRASADO",
            ).exists()

        if estudiante.ingreso_confirmado:
            estado_texto = "Presente"
            estado_clase = "presente"
        elif estudiante_atrasado:
            estado_texto = "Atrasado"
            estado_clase = "atrasado"
        else:
            estado_texto = "Pendiente"
            estado_clase = "pendiente"

        seguimiento_estudiantes.append({
            "nombre": estudiante.nombre_completo,
            "rut": estudiante.rut,
            "area": obtener_area_estudiante(estudiante),
            "plan": obtener_plan_estudiante(estudiante),
            "bloque": obtener_bloque_estudiante(estudiante),
            "estudiante_presente": estudiante.ingreso_confirmado,
            "estudiante_atrasado": estudiante_atrasado,
            "estado_texto": estado_texto,
            "estado_clase": estado_clase,
            "invitacion_1": invitacion_1.usada if invitacion_1 else False,
            "invitacion_1_atrasada": invitacion_1_atrasada,
            "invitacion_2": invitacion_2.usada if invitacion_2 else False,
            "invitacion_2_atrasada": invitacion_2_atrasada,
            "invitaciones_gestionadas": estudiante.invitaciones_entregadas,
        })

    ultimos_registros = registros_filtrados.order_by("-fecha_hora")[:30]

    registros = []
    for registro in ultimos_registros:
        estudiante = registro.estudiante
        registros.append({
            "nombre": estudiante.nombre_completo if estudiante else "Sin estudiante",
            "rut": estudiante.rut if estudiante else "-",
            "tipo_acceso": registro.tipo,
            "resultado": registro.get_resultado_display(),
            "codigo_resultado": registro.resultado,
            "ingreso_atrasado": registro.resultado == "ATRASADO",
            "hora": registro.fecha_hora.strftime("%H:%M:%S"),
            "area": obtener_area_estudiante(estudiante) if estudiante else "-",
            "plan": obtener_plan_estudiante(estudiante) if estudiante else "-",
            "bloque": obtener_bloque_estudiante(estudiante) if estudiante else "-",
        })

    registros_atrasados = registros_filtrados.filter(
        resultado="ATRASADO"
    ).order_by("-fecha_hora")[:80]

    atrasados = []
    for registro in registros_atrasados:
        estudiante = registro.estudiante
        atrasados.append({
            "nombre": estudiante.nombre_completo if estudiante else "Sin estudiante",
            "rut": estudiante.rut if estudiante else "-",
            "tipo_acceso": registro.tipo,
            "resultado": registro.get_resultado_display(),
            "hora": registro.fecha_hora.strftime("%H:%M:%S"),
            "area": obtener_area_estudiante(estudiante) if estudiante else "-",
            "plan": obtener_plan_estudiante(estudiante) if estudiante else "-",
            "bloque": obtener_bloque_estudiante(estudiante) if estudiante else "-",
            "ceremonia": (
                estudiante.bloque_ceremonia.ceremonia.nombre
                if estudiante
                and estudiante.bloque_ceremonia
                and estudiante.bloque_ceremonia.ceremonia
                else "-"
            ),
        })

    return JsonResponse({
        "kpis": {
            "estudiantes_totales": estudiantes_totales,
            "estudiantes_presentes": estudiantes_presentes,
            "estudiantes_pendientes": estudiantes_pendientes,
            "estudiantes_atrasados": estudiantes_atrasados,
            "invitados_presentes": invitados_presentes,
            "invitados_atrasados": invitados_atrasados,
            "invitaciones_totales": invitaciones_totales,
            "invitaciones_pendientes": invitaciones_pendientes,
            "total_asistentes": total_asistentes,
            "total_atrasados": total_atrasados,
            "alertas_qr": alertas_qr,
            "porcentaje_estudiantes": porcentaje_estudiantes,
            "porcentaje_invitaciones": porcentaje_invitaciones,
        },
        "avance_por_plan": avance_por_plan,
        "seguimiento_estudiantes": seguimiento_estudiantes,
        "ultimos_registros": registros,
        "ingresos_atrasados": atrasados,
        "busqueda": request.GET.get("q", "").strip(),
    })


def obtener_planes_por_area(request):
    area = request.GET.get("area", "").strip()

    planes = PlanEstudio.objects.select_related("area").all()

    if area:
        planes = planes.filter(area__nombre=area)

    planes = planes.values_list(
        "nombre", flat=True
    ).distinct().order_by("nombre")

    return JsonResponse({"planes": list(planes)})


def tarjetas_invitacion(request):
    busqueda = request.GET.get("q", "").strip()
    bloque_id = request.GET.get("bloque", "").strip()
    ceremonia_id = request.GET.get("ceremonia", "").strip()
    area = request.GET.get("area", "").strip()
    tipo = request.GET.get("tipo", "").strip()

    estudiantes = EstudianteTitulado.objects.select_related(
        "bloque_ceremonia",
        "bloque_ceremonia__ceremonia",
        "plan_estudio",
        "plan_estudio__area",
    ).prefetch_related("invitaciones").all().order_by(
        "nombre_completo",
        "bloque_ceremonia__fecha",
        "bloque_ceremonia__hora_inicio",
    )

    if busqueda:
        estudiantes = estudiantes.filter(
            Q(rut__icontains=busqueda) |
            Q(nombre_completo__icontains=busqueda)
        )

    if bloque_id:
        estudiantes = estudiantes.filter(bloque_ceremonia_id=bloque_id)

    if ceremonia_id:
        estudiantes = estudiantes.filter(bloque_ceremonia__ceremonia_id=ceremonia_id)

    if area:
        estudiantes = estudiantes.filter(plan_estudio__area__nombre=area)

    tarjetas = []

    for estudiante in estudiantes:
        asegurar_qr_estudiante(estudiante)
        datos_fecha = datos_fecha_bloque(estudiante.bloque_ceremonia)

        if tipo in ["", "ESTUDIANTE"]:
            tarjetas.append({
                "tipo": "ESTUDIANTE",
                "numero_invitacion": "",
                "nombre": estudiante.nombre_completo,
                "rut": estudiante.rut,
                "area": obtener_area_estudiante(estudiante),
                "plan": obtener_plan_estudiante(estudiante),
                "bloque": obtener_bloque_estudiante(estudiante),
                "dia_semana": datos_fecha["dia_semana"],
                "numero_dia": datos_fecha["numero_dia"],
                "mes": datos_fecha["mes_es"],
                "hora": datos_fecha["hora_es"],
                "lugar": DIRECCION_CEREMONIA,
                "qr": estudiante.imagen_qr_estudiante.url if estudiante.imagen_qr_estudiante else "",
                "codigo": estudiante.codigo_qr_estudiante,
                "utilizado": estudiante.ingreso_confirmado,
            })

        if tipo in ["", "INVITACION"]:
            for invitacion in estudiante.invitaciones.all().order_by("numero_invitacion"):
                tarjetas.append({
                    "tipo": "INVITACION",
                    "numero_invitacion": invitacion.numero_invitacion,
                    "nombre": estudiante.nombre_completo,
                    "rut": estudiante.rut,
                    "area": obtener_area_estudiante(estudiante),
                    "plan": obtener_plan_estudiante(estudiante),
                    "bloque": obtener_bloque_estudiante(estudiante),
                    "dia_semana": datos_fecha["dia_semana"],
                    "numero_dia": datos_fecha["numero_dia"],
                    "mes": datos_fecha["mes_es"],
                    "hora": datos_fecha["hora_es"],
                    "lugar": DIRECCION_CEREMONIA,
                    "qr": invitacion.imagen_qr.url if invitacion.imagen_qr else "",
                    "codigo": invitacion.codigo_qr,
                    "utilizado": invitacion.usada,
                })

    return render(
        request,
        "titulacion/tarjetas.html",
        {
            "tarjetas": tarjetas,
            "total_tarjetas": len(tarjetas),
            "bloques": BloqueCeremonia.objects.all().order_by("fecha", "hora_inicio"),
            "ceremonias": Ceremonia.objects.all().order_by("-anio", "nombre"),
            "areas": AreaAcademica.objects.all().order_by("nombre"),
            "busqueda": busqueda,
            "bloque_seleccionado": bloque_id,
            "ceremonia_seleccionada": ceremonia_id,
            "area_seleccionada": area,
            "tipo_seleccionado": tipo,
        }
    )


def entrega_invitaciones(request):
    busqueda = request.GET.get("q", "").strip()
    resultados = []

    if busqueda:
        estudiantes = EstudianteTitulado.objects.select_related(
            "bloque_ceremonia",
            "bloque_ceremonia__ceremonia",
            "plan_estudio",
            "plan_estudio__area",
        ).prefetch_related("invitaciones").filter(
            Q(rut__icontains=busqueda) |
            Q(nombre_completo__icontains=busqueda) |
            Q(plan_estudio__nombre__icontains=busqueda)
        ).order_by(
            "nombre_completo",
            "bloque_ceremonia__fecha",
            "bloque_ceremonia__hora_inicio",
            "plan_estudio__nombre",
        )

        for estudiante in estudiantes:
            asegurar_qr_estudiante(estudiante)
            datos_fecha = datos_fecha_bloque(estudiante.bloque_ceremonia)

            resultados.append({
                "estudiante": estudiante,
                "invitaciones": estudiante.invitaciones.all().order_by("numero_invitacion"),
                "fecha_es": datos_fecha["fecha_es"],
                "hora_es": datos_fecha["hora_es"],
                "dia_semana": datos_fecha["dia_semana"],
                "numero_dia": datos_fecha["numero_dia"],
                "mes_es": datos_fecha["mes_es"],
                "direccion_ceremonia": DIRECCION_CEREMONIA,
            })

    return render(
        request,
        "titulacion/entrega_invitaciones.html",
        {
            "busqueda": busqueda,
            "resultados": resultados,
            "total_resultados": len(resultados),
            "direccion_ceremonia": DIRECCION_CEREMONIA,
        }
    )


@require_POST
def marcar_entrega_invitaciones(request, estudiante_id):
    estudiante = get_object_or_404(EstudianteTitulado, id=estudiante_id)

    estudiante.invitaciones_entregadas = True
    estudiante.save()

    messages.success(
        request,
        f"Invitaciones de {estudiante.nombre_completo} para {estudiante.bloque_ceremonia.nombre} marcadas como gestionadas."
    )

    return redirect(f"/entrega-invitaciones/?q={estudiante.rut}")


def cambio_ceremonia(request):
    busqueda = request.GET.get("q", "").strip()
    estudiante_id = request.GET.get("estudiante", "").strip()

    estudiantes = EstudianteTitulado.objects.none()
    estudiante_seleccionado = None

    bloques_destino = BloqueCeremonia.objects.select_related(
        "ceremonia"
    ).order_by("fecha", "hora_inicio", "nombre")

    historial = CambioCeremonia.objects.select_related(
        "estudiante",
        "bloque_origen",
        "bloque_destino",
    ).order_by("-fecha_cambio")[:25]

    if busqueda:
        estudiantes = EstudianteTitulado.objects.select_related(
            "bloque_ceremonia",
            "bloque_ceremonia__ceremonia",
            "plan_estudio",
            "plan_estudio__area",
        ).prefetch_related("invitaciones").filter(
            Q(rut__icontains=busqueda) |
            Q(nombre_completo__icontains=busqueda) |
            Q(plan_estudio__nombre__icontains=busqueda)
        ).order_by(
            "nombre_completo",
            "bloque_ceremonia__fecha",
            "bloque_ceremonia__hora_inicio",
        )

    if estudiante_id:
        estudiante_seleccionado = get_object_or_404(
            EstudianteTitulado.objects.select_related(
                "bloque_ceremonia",
                "bloque_ceremonia__ceremonia",
                "plan_estudio",
                "plan_estudio__area",
            ).prefetch_related("invitaciones"),
            id=estudiante_id
        )

    if request.method == "POST":
        estudiante_id_post = request.POST.get("estudiante_id", "").strip()
        bloque_destino_id = request.POST.get("bloque_destino", "").strip()
        motivo = request.POST.get("motivo", "").strip()

        estudiante = get_object_or_404(
            EstudianteTitulado.objects.select_related(
                "bloque_ceremonia",
                "plan_estudio",
            ).prefetch_related("invitaciones"),
            id=estudiante_id_post
        )

        bloque_destino = get_object_or_404(BloqueCeremonia, id=bloque_destino_id)
        bloque_origen = estudiante.bloque_ceremonia

        if not motivo:
            messages.error(request, "Debes ingresar el motivo institucional del cambio.")
            return redirect(f"{request.path}?q={estudiante.rut}&estudiante={estudiante.id}")

        if bloque_destino.id == bloque_origen.id:
            messages.error(request, "El estudiante ya pertenece a la ceremonia seleccionada.")
            return redirect(f"{request.path}?q={estudiante.rut}&estudiante={estudiante.id}")

        if estudiante.ingreso_confirmado:
            messages.error(
                request,
                "No se puede cambiar de ceremonia porque el estudiante ya registra ingreso."
            )
            return redirect(f"{request.path}?q={estudiante.rut}&estudiante={estudiante.id}")

        if estudiante.invitaciones.filter(usada=True).exists():
            messages.error(
                request,
                "No se puede cambiar de ceremonia porque una o más invitaciones ya fueron utilizadas."
            )
            return redirect(f"{request.path}?q={estudiante.rut}&estudiante={estudiante.id}")

        duplicado = EstudianteTitulado.objects.filter(
            rut=estudiante.rut,
            bloque_ceremonia=bloque_destino,
            plan_estudio=estudiante.plan_estudio,
        ).exclude(id=estudiante.id).exists()

        if duplicado:
            messages.error(
                request,
                "No se puede realizar el cambio porque ya existe este RUT en el bloque destino con el mismo plan de estudio."
            )
            return redirect(f"{request.path}?q={estudiante.rut}&estudiante={estudiante.id}")

        usuario_responsable = None
        if request.user.is_authenticated:
            usuario_responsable = request.user.get_username()

        with transaction.atomic():
            CambioCeremonia.objects.create(
                estudiante=estudiante,
                bloque_origen=bloque_origen,
                bloque_destino=bloque_destino,
                motivo=motivo,
                usuario_responsable=usuario_responsable,
                observacion=(
                    "Cambio de ceremonia realizado desde módulo institucional. "
                    "Se regeneraron QR de estudiante e invitaciones."
                ),
            )

            RegistroIngreso.objects.create(
                estudiante=estudiante,
                invitacion=None,
                tipo="ESTUDIANTE",
                resultado="OTRA_CEREMONIA",
                observacion=(
                    f"CAMBIO ADMINISTRATIVO DE CEREMONIA: "
                    f"{bloque_origen.nombre} → {bloque_destino.nombre}. "
                    f"Motivo: {motivo}"
                ),
                usuario_registro=usuario_responsable,
            )

            estudiante.bloque_ceremonia = bloque_destino
            estudiante.invitaciones_entregadas = False
            estudiante.save()

            regenerar_qr_estudiante_e_invitaciones(estudiante)

        messages.success(
            request,
            f"El estudiante {estudiante.nombre_completo} fue cambiado correctamente a {bloque_destino.nombre}. Se regeneraron sus QR."
        )

        return redirect(f"{request.path}?q={estudiante.rut}&estudiante={estudiante.id}")

    return render(
        request,
        "titulacion/cambio_ceremonia.html",
        {
            "busqueda": busqueda,
            "estudiantes": estudiantes,
            "estudiante_seleccionado": estudiante_seleccionado,
            "bloques_destino": bloques_destino,
            "historial": historial,
        }
    )


@require_POST
def abrir_bloque_ceremonia(request, bloque_id):
    bloque = get_object_or_404(BloqueCeremonia, id=bloque_id)
    bloque.abrir()
    messages.success(request, f"La ceremonia {bloque.nombre} fue abierta correctamente.")
    return redirect("titulacion:registro_ingreso")


@require_POST
def cerrar_bloque_ceremonia(request, bloque_id):
    bloque = get_object_or_404(BloqueCeremonia, id=bloque_id)
    bloque.cerrar()
    messages.success(
        request,
        f"La ceremonia {bloque.nombre} fue cerrada correctamente. Los próximos ingresos quedarán como atrasados."
    )
    return redirect("titulacion:registro_ingreso")


@require_POST
def reprogramar_bloque_ceremonia(request, bloque_id):
    bloque = get_object_or_404(BloqueCeremonia, id=bloque_id)
    bloque.reprogramar()
    messages.success(request, f"La ceremonia {bloque.nombre} volvió al estado programada.")
    return redirect("titulacion:registro_ingreso")


@require_POST
def cerrar_registro_ceremonia(request, ceremonia_id):
    bloque = get_object_or_404(BloqueCeremonia, id=ceremonia_id)
    bloque.cerrar()
    messages.success(request, f"Registro de ingreso cerrado para {bloque.nombre}.")
    return redirect("titulacion:registro_ingreso")


@require_POST
def reabrir_registro_ceremonia(request, ceremonia_id):
    bloque = get_object_or_404(BloqueCeremonia, id=ceremonia_id)
    bloque.abrir()
    messages.success(request, f"Registro de ingreso reabierto para {bloque.nombre}.")
    return redirect("titulacion:registro_ingreso")


@require_POST
def cerrar_registro_bloque(request, bloque_id):
    bloque = get_object_or_404(BloqueCeremonia, id=bloque_id)
    bloque.cerrar()
    messages.success(request, f"Registro cerrado para {bloque.nombre}.")
    return redirect("titulacion:panel_control")


@require_POST
def reabrir_registro_bloque(request, bloque_id):
    bloque = get_object_or_404(BloqueCeremonia, id=bloque_id)
    bloque.abrir()
    messages.success(request, f"Registro reabierto para {bloque.nombre}.")
    return redirect("titulacion:panel_control")


def descargar_invitacion_pdf(request, tipo, estudiante_id):
    return redirect(f"/entrega-invitaciones/?q={estudiante_id}")


def api_dashboard(request):
    estudiantes = EstudianteTitulado.objects.all()

    return JsonResponse({
        "total_estudiantes": estudiantes.count(),
        "presentes": estudiantes.filter(ingreso_confirmado=True).count(),
        "pendientes": estudiantes.filter(ingreso_confirmado=False).count(),
        "atrasados": RegistroIngreso.objects.filter(resultado="ATRASADO").count(),
        "invitados_presentes": Invitacion.objects.filter(usada=True).count(),
        "invitados_atrasados": RegistroIngreso.objects.filter(
            tipo="INVITADO",
            resultado="ATRASADO"
        ).count(),
        "timestamp": timezone.now().strftime("%H:%M:%S"),
    })


def agregar_estudiante(request):
    bloques = BloqueCeremonia.objects.select_related(
        "ceremonia"
    ).order_by("ceremonia__nombre", "fecha", "hora_inicio", "nombre")

    areas = AreaAcademica.objects.order_by("nombre")

    planes = PlanEstudio.objects.select_related("area").order_by(
        "area__nombre", "nombre"
    )

    planes_json = json.dumps([
        {
            "id": p.id,
            "nombre": p.nombre,
            "area": p.area.nombre,
            "institucion": p.institucion or "",
        }
        for p in planes
    ])

    ultimos = EstudianteTitulado.objects.select_related(
        "bloque_ceremonia__ceremonia",
        "plan_estudio__area",
    ).order_by("-fecha_creacion")[:10]

    if request.method == "POST":
        rut = request.POST.get("rut", "").strip()
        nombre_completo = request.POST.get("nombre_completo", "").strip()
        bloque_id = request.POST.get("bloque_id", "").strip()
        plan_id = request.POST.get("plan_id", "").strip()
        jornada = request.POST.get("jornada", "").strip() or None
        correo = request.POST.get("correo", "").strip() or None
        telefono = request.POST.get("telefono", "").strip() or None

        errores = []

        if not rut:
            errores.append("El RUT es obligatorio.")
        if not nombre_completo:
            errores.append("El nombre completo es obligatorio.")
        if not bloque_id:
            errores.append("Debe seleccionar un bloque de ceremonia.")
        if not plan_id:
            errores.append("Debe seleccionar un plan de estudio.")

        bloque = None
        plan = None

        if not errores:
            try:
                bloque = BloqueCeremonia.objects.select_related("ceremonia").get(id=bloque_id)
            except BloqueCeremonia.DoesNotExist:
                errores.append("El bloque de ceremonia seleccionado no existe.")

            try:
                plan = PlanEstudio.objects.select_related("area").get(id=plan_id)
            except PlanEstudio.DoesNotExist:
                errores.append("El plan de estudio seleccionado no existe.")

        if not errores:
            duplicado = EstudianteTitulado.objects.filter(
                rut=rut,
                bloque_ceremonia=bloque,
                plan_estudio=plan,
            ).exists()

            if duplicado:
                errores.append(
                    f"El estudiante con RUT {rut} ya está registrado en el bloque "
                    f"'{bloque.nombre}' con el plan '{plan.nombre}'. No se puede duplicar."
                )

        if not errores:
            try:
                with transaction.atomic():
                    estudiante = EstudianteTitulado.objects.create(
                        rut=rut,
                        nombre_completo=nombre_completo,
                        bloque_ceremonia=bloque,
                        plan_estudio=plan,
                        jornada=jornada,
                        correo=correo,
                        telefono=telefono,
                    )
                    generar_qr_estudiante(estudiante)
                    crear_invitaciones_para_estudiante(estudiante)

                messages.success(
                    request,
                    f"Estudiante '{nombre_completo}' (RUT: {rut}) agregado correctamente "
                    f"al bloque '{bloque.nombre}'. Se generaron 2 invitaciones con QR."
                )
                return redirect("titulacion:agregar_estudiante")

            except IntegrityError:
                errores.append(
                    f"No se pudo guardar el estudiante con RUT {rut}. "
                    "Verifique que no exista un registro duplicado."
                )

        for error in errores:
            messages.error(request, error)

    return render(
        request,
        "titulacion/agregar_estudiante.html",
        {
            "bloques": bloques,
            "areas": areas,
            "planes_json": planes_json,
            "ultimos": ultimos,
        }
    )


def _calcular_reportes(bloque_id=None):
    est_qs = EstudianteTitulado.objects.select_related(
        "bloque_ceremonia", "bloque_ceremonia__ceremonia",
        "plan_estudio", "plan_estudio__area",
    )
    reg_qs = RegistroIngreso.objects.select_related(
        "estudiante", "estudiante__bloque_ceremonia",
        "invitacion", "invitacion__estudiante",
    )
    inv_qs = Invitacion.objects.select_related(
        "estudiante", "estudiante__bloque_ceremonia"
    )

    if bloque_id:
        est_qs = est_qs.filter(bloque_ceremonia_id=bloque_id)
        reg_qs = reg_qs.filter(estudiante__bloque_ceremonia_id=bloque_id)
        inv_qs = inv_qs.filter(estudiante__bloque_ceremonia_id=bloque_id)

    total_titulados = est_qs.count()
    total_ingresados = est_qs.filter(ingreso_confirmado=True).count()
    total_ausentes = total_titulados - total_ingresados
    total_inv_usadas = inv_qs.filter(usada=True).count()
    total_personas = total_ingresados + total_inv_usadas
    pct_asistencia = round(total_ingresados / total_titulados * 100, 1) if total_titulados > 0 else 0.0

    hora_peak_row = (
        reg_qs
        .filter(resultado__in=["PERMITIDO", "ATRASADO"])
        .annotate(hora=TruncHour("fecha_hora"))
        .values("hora")
        .annotate(n=Count("id"))
        .order_by("-n")
        .first()
    )
    hora_peak = hora_peak_row["hora"].strftime("%H:00") if hora_peak_row and hora_peak_row["hora"] else "-"

    ultimo_reg = reg_qs.filter(resultado__in=["PERMITIDO", "ATRASADO"]).order_by("-fecha_hora").first()
    ultimo_ingreso_str = ultimo_reg.fecha_hora.strftime("%d/%m %H:%M") if ultimo_reg else "-"

    mejor = (
        EstudianteTitulado.objects
        .values("bloque_ceremonia__nombre")
        .annotate(n=Count("id", filter=Q(ingreso_confirmado=True)))
        .order_by("-n")
        .first()
    )
    ceremonia_mayor_asistencia = mejor["bloque_ceremonia__nombre"] if mejor else "-"

    peor = (
        RegistroIngreso.objects
        .filter(resultado="ATRASADO")
        .values("estudiante__bloque_ceremonia__nombre")
        .annotate(n=Count("id"))
        .order_by("-n")
        .first()
    )
    ceremonia_mayor_atraso = (
        peor["estudiante__bloque_ceremonia__nombre"]
        if peor and peor["estudiante__bloque_ceremonia__nombre"]
        else "-"
    )

    dashboard = {
        "total_titulados": total_titulados,
        "total_ingresados": total_ingresados,
        "total_ausentes": total_ausentes,
        "total_inv_usadas": total_inv_usadas,
        "total_personas": total_personas,
        "pct_asistencia": pct_asistencia,
        "ceremonia_mayor_asistencia": ceremonia_mayor_asistencia,
        "ceremonia_mayor_atraso": ceremonia_mayor_atraso,
        "ultimo_ingreso": ultimo_ingreso_str,
        "hora_peak": hora_peak,
    }

    bloques_qs = BloqueCeremonia.objects.select_related("ceremonia").order_by("fecha", "hora_inicio")
    if bloque_id:
        bloques_qs = bloques_qs.filter(id=bloque_id)

    por_bloque = []
    for bloque in bloques_qs:
        est_b = EstudianteTitulado.objects.filter(bloque_ceremonia=bloque)
        esperados = est_b.count()
        ingresados_b = est_b.filter(ingreso_confirmado=True).count()
        ausentes_b = esperados - ingresados_b
        pct_b = round(ingresados_b / esperados * 100, 1) if esperados > 0 else 0.0

        inv_b = Invitacion.objects.filter(estudiante__bloque_ceremonia=bloque)
        inv_usadas_b = inv_b.filter(usada=True).count()
        inv_no_usadas_b = inv_b.filter(usada=False).count()
        total_asistentes_b = ingresados_b + inv_usadas_b

        reg_b = RegistroIngreso.objects.filter(estudiante__bloque_ceremonia=bloque)
        atrasados_b = reg_b.filter(resultado="ATRASADO").values("estudiante_id").distinct().count()

        tiempos_b = reg_b.filter(resultado__in=["PERMITIDO", "ATRASADO"]).aggregate(
            primer=Min("fecha_hora"), ultimo=Max("fecha_hora")
        )
        primer_b = tiempos_b["primer"].strftime("%H:%M") if tiempos_b["primer"] else "-"
        ultimo_b = tiempos_b["ultimo"].strftime("%H:%M") if tiempos_b["ultimo"] else "-"

        tiempo_flujo_b = "-"
        if tiempos_b["primer"] and tiempos_b["ultimo"]:
            delta = tiempos_b["ultimo"] - tiempos_b["primer"]
            mins = int(delta.total_seconds() / 60)
            h, m = divmod(mins, 60)
            tiempo_flujo_b = f"{h}h {m}min" if h > 0 else f"{m} min"

        hora_peak_b_row = (
            reg_b
            .filter(resultado__in=["PERMITIDO", "ATRASADO"])
            .annotate(hora=TruncHour("fecha_hora"))
            .values("hora")
            .annotate(n=Count("id"))
            .order_by("-n")
            .first()
        )
        hora_peak_b = hora_peak_b_row["hora"].strftime("%H:00") if hora_peak_b_row and hora_peak_b_row["hora"] else "-"

        por_bloque.append({
            "bloque": bloque.nombre,
            "fecha": bloque.fecha.strftime("%d/%m/%Y"),
            "hora_inicio": str(bloque.hora_inicio)[:5],
            "estado": bloque.estado_registro,
            "esperados": esperados,
            "ingresados": ingresados_b,
            "ausentes": ausentes_b,
            "pct": pct_b,
            "inv_usadas": inv_usadas_b,
            "inv_no_usadas": inv_no_usadas_b,
            "total_asistentes": total_asistentes_b,
            "atrasados": atrasados_b,
            "primer_ingreso": primer_b,
            "ultimo_ingreso": ultimo_b,
            "tiempo_flujo": tiempo_flujo_b,
            "hora_peak": hora_peak_b,
        })

    total_reg = reg_qs.count()
    qr_validos = reg_qs.filter(resultado="PERMITIDO").count()
    qr_invalidos = reg_qs.filter(resultado__in=["DENEGADO", "OTRA_CEREMONIA"]).count()
    qr_duplicados = reg_qs.filter(resultado="DUPLICADO").count()
    qr_atrasados = reg_qs.filter(resultado="ATRASADO").count()

    por_operador = list(
        reg_qs.values("usuario_registro").annotate(n=Count("id")).order_by("-n")
    )

    operativo = {
        "total_registros": total_reg,
        "qr_validos": qr_validos,
        "qr_invalidos": qr_invalidos,
        "qr_duplicados": qr_duplicados,
        "qr_atrasados": qr_atrasados,
        "intentos_fallidos": qr_invalidos + qr_duplicados,
        "por_operador": [
            {"usuario": o["usuario_registro"] or "Sistema", "total": o["n"]}
            for o in por_operador
        ],
    }

    por_hora = list(
        reg_qs
        .filter(resultado__in=["PERMITIDO", "ATRASADO"])
        .annotate(hora=TruncHour("fecha_hora"))
        .values("hora")
        .annotate(n=Count("id"))
        .order_by("hora")
    )
    por_hora_fmt = [
        {"hora": item["hora"].strftime("%H:00") if item["hora"] else "?", "total": item["n"]}
        for item in por_hora
    ]

    tiempos_global = reg_qs.filter(resultado__in=["PERMITIDO", "ATRASADO"]).aggregate(
        primer=Min("fecha_hora"), ultimo=Max("fecha_hora"), total=Count("id")
    )
    prom_por_min = "-"
    if tiempos_global["primer"] and tiempos_global["ultimo"] and tiempos_global["total"]:
        duracion_min = max(1, int((tiempos_global["ultimo"] - tiempos_global["primer"]).total_seconds() / 60))
        prom_por_min = round(tiempos_global["total"] / duracion_min, 2)

    tiempos = {
        "por_hora": por_hora_fmt,
        "hora_peak": hora_peak,
        "prom_por_minuto": prom_por_min,
    }

    return {
        "dashboard": dashboard,
        "por_bloque": por_bloque,
        "operativo": operativo,
        "tiempos": tiempos,
    }


def reportes(request):
    bloques = BloqueCeremonia.objects.select_related("ceremonia").order_by("fecha", "hora_inicio", "nombre")
    return render(request, "titulacion/reportes.html", {"bloques": bloques})


def datos_reportes(request):
    bloque_id = request.GET.get("bloque", "").strip() or None
    return JsonResponse(_calcular_reportes(bloque_id=bloque_id))


def exportar_reportes_excel(request):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    bloque_id = request.GET.get("bloque", "").strip() or None
    datos = _calcular_reportes(bloque_id=bloque_id)

    wb = openpyxl.Workbook()

    def eh(cell, bg="06152F"):
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def cw(ws, col, w):
        ws.column_dimensions[get_column_letter(col)].width = w

    # --- Hoja 1: Dashboard ---
    ws1 = wb.active
    ws1.title = "Dashboard"
    ws1.merge_cells("A1:B1")
    c = ws1["A1"]
    c.value = "REPORTE EJECUTIVO — CEREMONIA TITULACIÓN INACAP 2026"
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = PatternFill("solid", fgColor="E30613")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 22
    ws1.append(["Métrica", "Valor"])
    for i in range(1, 3):
        eh(ws1.cell(ws1.max_row, i))
    d = datos["dashboard"]
    for nombre, valor in [
        ("Total titulados cargados", d["total_titulados"]),
        ("Total titulados ingresados", d["total_ingresados"]),
        ("Total ausentes", d["total_ausentes"]),
        ("Total invitados registrados", d["total_inv_usadas"]),
        ("Total personas ingresadas", d["total_personas"]),
        ("% Asistencia general", f"{d['pct_asistencia']}%"),
        ("Ceremonia mayor asistencia", d["ceremonia_mayor_asistencia"]),
        ("Ceremonia mayor atraso", d["ceremonia_mayor_atraso"]),
        ("Último ingreso registrado", d["ultimo_ingreso"]),
        ("Hora peak de ingresos", d["hora_peak"]),
    ]:
        ws1.append([nombre, valor])
    cw(ws1, 1, 42)
    cw(ws1, 2, 28)

    # --- Hoja 2: Por Ceremonia ---
    ws2 = wb.create_sheet("Por Ceremonia")
    cols2 = ["Bloque", "Fecha", "Hora", "Estado", "Esperados", "Ingresados",
             "Ausentes", "% Asist.", "Inv. Usadas", "Inv. No usadas",
             "Total asistentes", "Atrasados", "Primer ingreso", "Último ingreso",
             "Tiempo flujo", "Hora peak"]
    ws2.append(cols2)
    for i in range(1, len(cols2) + 1):
        eh(ws2.cell(1, i))
    for b in datos["por_bloque"]:
        ws2.append([
            b["bloque"], b["fecha"], b["hora_inicio"], b["estado"],
            b["esperados"], b["ingresados"], b["ausentes"], f"{b['pct']}%",
            b["inv_usadas"], b["inv_no_usadas"], b["total_asistentes"],
            b["atrasados"], b["primer_ingreso"], b["ultimo_ingreso"],
            b["tiempo_flujo"], b["hora_peak"],
        ])
    for i, w in enumerate([30, 12, 8, 12, 10, 10, 10, 10, 11, 13, 14, 10, 13, 13, 12, 10], start=1):
        cw(ws2, i, w)

    # --- Hoja 3: Operativo ---
    ws3 = wb.create_sheet("Operativo")
    ws3.append(["Métrica QR", "Total"])
    for i in range(1, 3):
        eh(ws3.cell(1, i))
    op = datos["operativo"]
    for nombre, valor in [
        ("QR válidos (PERMITIDO)", op["qr_validos"]),
        ("QR inválidos (DENEGADO + OTRA_CEREMONIA)", op["qr_invalidos"]),
        ("QR duplicados", op["qr_duplicados"]),
        ("Ingresos atrasados", op["qr_atrasados"]),
        ("Intentos fallidos", op["intentos_fallidos"]),
        ("Total registros", op["total_registros"]),
    ]:
        ws3.append([nombre, valor])
    ws3.append([])
    ws3.append(["Operador", "Registros"])
    for i in range(1, 3):
        eh(ws3.cell(ws3.max_row, i))
    for o in op["por_operador"]:
        ws3.append([o["usuario"], o["total"]])
    cw(ws3, 1, 40)
    cw(ws3, 2, 14)

    # --- Hoja 4: Tiempos ---
    ws4 = wb.create_sheet("Tiempos")
    ws4.append(["Hora", "Ingresos"])
    for i in range(1, 3):
        eh(ws4.cell(1, i))
    for item in datos["tiempos"]["por_hora"]:
        ws4.append([item["hora"], item["total"]])
    ws4.append([])
    ws4.append(["Promedio registros/min", datos["tiempos"]["prom_por_minuto"]])
    cw(ws4, 1, 14)
    cw(ws4, 2, 14)

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="reporte_titulacion_2026.xlsx"'
    wb.save(resp)
    return resp


def healthcheck(request):
    return HttpResponse("OK")